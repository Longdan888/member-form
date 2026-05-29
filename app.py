"""
在线会员信息收集工具
支持本地运行和云端部署，SQLite 数据库 + Excel 导出。
"""
import os
import re
import sqlite3
from datetime import datetime
from functools import wraps

from flask import Flask, request, render_template, send_file, redirect, url_for, flash, session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from werkzeug.middleware.proxy_fix import ProxyFix

# ---- Flask 初始化 ----
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production-2026")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

# ---- 配置 ----
DB_PATH = os.path.join(os.path.dirname(__file__), "data", "submissions.db")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

HEADERS = ["提交时间", "姓名", "手机号", "会员类型", "会员账号/ID", "备注"]
MEMBERSHIP_TYPES = ["88VIP", "京东Plus", "i茅台", "龙蛋业务", "AD钙", "其他"]


# ======================================================================
#  SQLite 数据库操作
# ======================================================================
def get_db():
    """获取数据库连接（每个线程独立连接）"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")       # 读写并发优化
    conn.execute("PRAGMA busy_timeout=5000")       # 忙等待 5 秒
    return conn


def init_db():
    """初始化数据库：建表 + 尝试从旧 Excel 迁移数据"""
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS submissions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at  TEXT    NOT NULL,
            name        TEXT    NOT NULL,
            phone       TEXT    NOT NULL UNIQUE,
            membership_type TEXT NOT NULL,
            account_id  TEXT    DEFAULT '',
            remarks     TEXT    DEFAULT ''
        )
    """)
    conn.commit()
    _migrate_from_excel(conn)
    conn.close()


def _migrate_from_excel(conn):
    """如果数据库为空且旧 Excel 存在，自动迁移数据"""
    count = conn.execute("SELECT COUNT(*) FROM submissions").fetchone()[0]
    if count > 0:
        return  # 已有数据，跳过

    excel_path = os.path.join(os.path.dirname(__file__), "data", "submissions.xlsx")
    if not os.path.exists(excel_path):
        return

    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    migrated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            conn.execute(
                "INSERT OR IGNORE INTO submissions (created_at, name, phone, membership_type, account_id, remarks) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (str(row[0] or ""), str(row[1] or ""), str(row[2] or ""),
                 str(row[3] or ""), str(row[4] or ""), str(row[5] or ""))
            )
            migrated += 1
        except Exception:
            pass
    conn.commit()
    if migrated:
        print(f"[迁移] 从 Excel 导入了 {migrated} 条历史数据")


def insert_submission(data: dict) -> bool:
    """插入一条记录，手机号重复返回 False"""
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO submissions (created_at, name, phone, membership_type, account_id, remarks) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                data["name"], data["phone"], data["membership_type"],
                data.get("account_id", ""), data.get("remarks", ""),
            )
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False  # 手机号重复
    finally:
        conn.close()


def check_duplicate(phone: str) -> bool:
    """检查手机号是否已存在"""
    conn = get_db()
    row = conn.execute("SELECT 1 FROM submissions WHERE phone = ?", (phone,)).fetchone()
    conn.close()
    return row is not None


def get_all_rows():
    """获取全部记录，返回 [列表] 格式（兼容原模板）"""
    conn = get_db()
    rows = conn.execute(
        "SELECT created_at, name, phone, membership_type, account_id, remarks "
        "FROM submissions ORDER BY id DESC"
    ).fetchall()
    conn.close()
    return [list(r) for r in rows]


def get_stats():
    """获取各会员类型统计"""
    conn = get_db()
    rows = conn.execute(
        "SELECT membership_type, COUNT(*) as cnt FROM submissions "
        "GROUP BY membership_type ORDER BY cnt DESC"
    ).fetchall()
    conn.close()
    return {r["membership_type"]: r["cnt"] for r in rows}


def get_total():
    """获取总记录数"""
    conn = get_db()
    row = conn.execute("SELECT COUNT(*) FROM submissions").fetchone()
    conn.close()
    return row[0]


def get_rows_by_type(membership_type: str):
    """按会员类型筛选记录"""
    conn = get_db()
    rows = conn.execute(
        "SELECT created_at, name, phone, membership_type, account_id, remarks "
        "FROM submissions WHERE membership_type = ? ORDER BY id DESC",
        (membership_type,)
    ).fetchall()
    conn.close()
    return [list(r) for r in rows]


# ======================================================================
#  Excel 导出工具
# ======================================================================
def build_excel(rows):
    """根据数据行生成 Excel Workbook，rows 为列表的列表（不含表头）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "会员信息"
    ws.append(HEADERS)

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    col_widths = [20, 12, 16, 14, 18, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row_data in rows:
        ws.append(row_data)

    return wb


# ======================================================================
#  验证
# ======================================================================
def validate_phone(phone: str) -> bool:
    return bool(re.match(r"^1[3-9]\d{9}$", phone))


# ======================================================================
#  管理后台认证装饰器
# ======================================================================
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            flash("请先登录管理后台", "warning")
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


# ======================================================================
#  路由 — 用户表单
# ======================================================================
@app.route("/")
@app.route("/form")
def form():
    return render_template("form.html", membership_types=MEMBERSHIP_TYPES)


@app.route("/submit", methods=["POST"])
def submit():
    name = request.form.get("name", "").strip()
    phone = request.form.get("phone", "").strip()
    membership_type = request.form.get("membership_type", "").strip()
    account_id = request.form.get("account_id", "").strip()
    remarks = request.form.get("remarks", "").strip()

    # 后端校验
    errors = []
    if not name:
        errors.append("请输入姓名")
    if not phone:
        errors.append("请输入手机号")
    elif not validate_phone(phone):
        errors.append("手机号格式不正确")
    if membership_type not in MEMBERSHIP_TYPES:
        errors.append("请选择有效的会员类型")

    if errors:
        return render_template(
            "form.html", membership_types=MEMBERSHIP_TYPES, errors=errors,
            form_data={"name": name, "phone": phone, "membership_type": membership_type,
                       "account_id": account_id, "remarks": remarks},
        )

    # 检查重复
    if check_duplicate(phone):
        return render_template(
            "form.html", membership_types=MEMBERSHIP_TYPES,
            errors=["该手机号已提交过，请勿重复提交"],
            form_data={"name": name, "phone": phone, "membership_type": membership_type,
                       "account_id": account_id, "remarks": remarks},
        )

    # 写入数据库
    insert_submission({
        "name": name, "phone": phone, "membership_type": membership_type,
        "account_id": account_id, "remarks": remarks,
    })

    return render_template("success.html", name=name)


# ======================================================================
#  路由 — 管理后台
# ======================================================================
@app.route("/admin")
@admin_required
def admin():
    rows = get_all_rows()
    stats = get_stats()
    total = get_total()
    # 模板需要表头+数据行格式
    return render_template("admin.html", headers=HEADERS, rows=rows, stats=stats, total=total)


@app.route("/admin/download")
@admin_required
def download():
    """下载全部数据为 Excel"""
    rows = get_all_rows()
    wb = build_excel(rows)
    export_path = os.path.join(os.path.dirname(__file__), "data", "_export_all.xlsx")
    wb.save(export_path)
    return send_file(
        export_path, as_attachment=True,
        download_name=f"会员信息_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/export/<membership_type>")
@admin_required
def export_by_type(membership_type):
    """按会员类型导出 Excel"""
    if membership_type not in MEMBERSHIP_TYPES:
        return "无效的会员类型", 400

    rows = get_rows_by_type(membership_type)
    if not rows:
        return f"没有找到 {membership_type} 类型的记录", 404

    wb = build_excel(rows)
    wb.active.title = membership_type
    export_path = os.path.join(os.path.dirname(__file__), "data", f"_export_{membership_type}.xlsx")
    wb.save(export_path)
    return send_file(
        export_path, as_attachment=True,
        download_name=f"{membership_type}_会员_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ======================================================================
#  路由 — 登录 / 退出
# ======================================================================
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("password", "") == ADMIN_PASSWORD:
            session["admin_logged_in"] = True
            flash("登录成功", "success")
            return redirect(url_for("admin"))
        else:
            flash("密码错误，请重试", "error")
    return render_template("login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    flash("已退出登录", "info")
    return redirect(url_for("admin_login"))


# ======================================================================
#  错误处理
# ======================================================================
@app.errorhandler(404)
def not_found(e):
    return render_template("form.html", membership_types=MEMBERSHIP_TYPES)


# ======================================================================
#  启动
# ======================================================================
def print_banner():
    print("=" * 50)
    print("  在线会员信息收集工具  v2.0")
    print("  数据库: SQLite (data/submissions.db)")
    print("  登记表单: http://localhost:5000/form")
    print("  管理后台: http://localhost:5000/admin")
    print("  管理密码:", ADMIN_PASSWORD)
    print("=" * 50)


if __name__ == "__main__":
    import sys
    init_db()

    if "--prod" in sys.argv:
        # 生产模式（云端部署）
        from waitress import serve
        print_banner()
        port = int(os.environ.get("PORT", 5000))
        print(f"  运行模式: 生产 (waitress) 端口:{port}")
        print("=" * 50)
        serve(app, host="0.0.0.0", port=port)
    else:
        # 开发模式（本地调试）
        print_banner()
        print("  运行模式: 开发 (Flask debug)")
        print("=" * 50)
        app.run(host="0.0.0.0", port=5000, debug=True)
